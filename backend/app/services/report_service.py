import csv
import io
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Dict, Any
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, desc, asc

import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.models.user import User
from app.models.student import Student
from app.models.loan import Loan
from app.models.inventory import Item
from app.models.program import Program
from app.models.activity import Activity
from app.models.activity_attendance import ActivityAttendance

class ReportService:
    @staticmethod
    def build_students_query(
        db: Session,
        skip: int = 0,
        limit: int = 100,
        search: Optional[str] = None,
        program_id: Optional[int] = None,
        status: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        sort_by: str = "created_at",
        order: str = "desc"
    ):
        query = db.query(Student).join(User, Student.user_id == User.id).outerjoin(Program, Student.program_id == Program.id)

        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                or_(
                    User.first_name.ilike(search_pattern),
                    User.last_name.ilike(search_pattern),
                    Student.national_id.ilike(search_pattern),
                    User.email.ilike(search_pattern)
                )
            )

        if program_id:
            query = query.filter(Student.program_id == program_id)
            
        if status and status != "ALL":
            query = query.filter(User.status == status)

        if date_from:
            try:
                dt_from = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc)
                query = query.filter(Student.created_at >= dt_from)
            except ValueError:
                pass
                
        if date_to:
            try:
                dt_to = datetime.fromisoformat(date_to).replace(tzinfo=timezone.utc) + timedelta(days=1)
                query = query.filter(Student.created_at < dt_to)
            except ValueError:
                pass

        sort_map = {
            "created_at": Student.created_at,
            "first_name": User.first_name,
            "last_name": User.last_name,
            "national_id": Student.national_id,
            "social_hours_completed": Student.social_hours_completed,
            "status": User.status
        }
        sort_attr = sort_map.get(sort_by, Student.created_at)

        if order == "asc":
            query = query.order_by(sort_attr.asc())
        else:
            query = query.order_by(sort_attr.desc())

        total = query.count()
        students = query.offset(skip).limit(limit).all()
        
        result_items = []
        for s in students:
            result_items.append({
                "id": s.id,
                "first_name": s.user.first_name if s.user else "",
                "last_name": s.user.last_name if s.user else "",
                "national_id": s.national_id,
                "email": s.user.email if s.user else "",
                "phone": s.user.phone if s.user else "",
                "program_name": s.program.name if s.program else "Sin Programa",
                "social_hours_completed": s.social_hours_completed,
                "status": str(s.user.status) if s.user else "",
                "created_at": s.created_at.isoformat() if s.created_at else ""
            })

        return {"items": result_items, "total": total}

    @staticmethod
    def build_loans_query(
        db: Session,
        skip: int = 0,
        limit: int = 100,
        search: Optional[str] = None,
        program_id: Optional[int] = None,
        status: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        sort_by: str = "created_at",
        order: str = "desc"
    ):
        query = db.query(Loan).join(Student).join(User, Student.user_id == User.id).join(Item).outerjoin(Program, Student.program_id == Program.id)

        if status and status != "ALL":
            query = query.filter(Loan.status == status)
            
        if program_id:
            query = query.filter(Student.program_id == program_id)
            
        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                or_(
                    User.first_name.ilike(search_pattern),
                    User.last_name.ilike(search_pattern),
                    Student.national_id.ilike(search_pattern),
                    Item.name.ilike(search_pattern)
                )
            )

        if date_from:
            try:
                dt_from = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc)
                query = query.filter(Loan.created_at >= dt_from)
            except ValueError:
                pass
                
        if date_to:
            try:
                dt_to = datetime.fromisoformat(date_to).replace(tzinfo=timezone.utc) + timedelta(days=1)
                query = query.filter(Loan.created_at < dt_to)
            except ValueError:
                pass

        sort_map = {
            "created_at": Loan.created_at,
            "start_time": Loan.start_time,
            "status": Loan.status,
            "hours_earned": Loan.hours_earned
        }
        sort_attr = sort_map.get(sort_by, Loan.created_at)

        if order == "asc":
            query = query.order_by(sort_attr.asc())
        else:
            query = query.order_by(sort_attr.desc())

        total = query.count()
        loans = query.offset(skip).limit(limit).all()
        
        result_items = []
        for l in loans:
            student = l.student
            user = student.user if student else None
            issuer = l.issuer
            
            result_items.append({
                "id": l.id,
                "first_name": user.first_name if user else "",
                "last_name": user.last_name if user else "",
                "national_id": student.national_id if student else "",
                "program_name": student.program.name if student and student.program else "Sin Programa",
                "item_name": l.item.name if l.item else "",
                "status": l.status,
                "start_time": l.start_time.isoformat() if l.start_time else "",
                "returned_time": l.returned_time.isoformat() if l.returned_time else "",
                "hours_earned": l.hours_earned,
                "issuer_name": f"{issuer.first_name} {issuer.last_name}" if issuer else "",
                "created_at": l.created_at.isoformat() if l.created_at else ""
            })

        return {"items": result_items, "total": total}

    @staticmethod
    def build_activities_query(
        db: Session,
        skip: int = 0,
        limit: int = 100,
        search: Optional[str] = None,
        program_id: Optional[int] = None,
        status: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        activity_id: Optional[int] = None,
        sort_by: str = "created_at",
        order: str = "desc"
    ):
        query = db.query(ActivityAttendance).join(Activity).join(Student).join(User, Student.user_id == User.id).outerjoin(Program, Student.program_id == Program.id)

        if status and status != "ALL":
            # For activities, status might refer to activity status or attendance status (if any)
            # User wants report of students who attended, so we filter by activity status if provided
            query = query.filter(Activity.status == status)
            
        if program_id:
            query = query.filter(Student.program_id == program_id)
            
        if activity_id:
            query = query.filter(Activity.id == activity_id)
            
        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                or_(
                    User.first_name.ilike(search_pattern),
                    User.last_name.ilike(search_pattern),
                    Student.national_id.ilike(search_pattern),
                    Activity.title.ilike(search_pattern)
                )
            )

        if date_from:
            try:
                dt_from = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc)
                query = query.filter(ActivityAttendance.scanned_at >= dt_from)
            except ValueError:
                pass
                
        if date_to:
            try:
                dt_to = datetime.fromisoformat(date_to).replace(tzinfo=timezone.utc) + timedelta(days=1)
                query = query.filter(ActivityAttendance.scanned_at < dt_to)
            except ValueError:
                pass

        sort_map = {
            "created_at": ActivityAttendance.scanned_at,
            "scanned_at": ActivityAttendance.scanned_at,
            "title": Activity.title,
            "hours_earned": ActivityAttendance.hours_earned
        }
        sort_attr = sort_map.get(sort_by, ActivityAttendance.scanned_at)

        if order == "asc":
            query = query.order_by(sort_attr.asc())
        else:
            query = query.order_by(sort_attr.desc())

        total = query.count()
        attendances = query.offset(skip).limit(limit).all()
        
        result_items = []
        for att in attendances:
            student = att.student
            user = student.user if student else None
            activity = att.activity
            
            result_items.append({
                "id": att.id,
                "activity_title": activity.title if activity else "",
                "first_name": user.first_name if user else "",
                "last_name": user.last_name if user else "",
                "national_id": student.national_id if student else "",
                "program_name": student.program.name if student and student.program else "Sin Programa",
                "hours_earned": att.hours_earned,
                "scanned_at": att.scanned_at.isoformat() if att.scanned_at else "",
                "event_datetime": activity.event_datetime.isoformat() if activity and activity.event_datetime else "",
                "location": activity.location if activity else "",
                "hours_reward": activity.hours_reward if activity else 0,
                "activity_status": activity.status if activity else ""
            })

        return {"items": result_items, "total": total}

    @staticmethod
    def export_csv(data: List[Dict[str, Any]], fields: List[str], field_labels: Dict[str, str]):
        stream = io.StringIO()
        writer = csv.writer(stream)
        
        headers = [field_labels.get(f, f) for f in fields]
        writer.writerow(headers)
        
        for row in data:
            writer.writerow([str(row.get(f, "")) for f in fields])
            
        response = StreamingResponse(iter([stream.getvalue()]), media_type="text/csv")
        response.headers["Content-Disposition"] = "attachment; filename=report.csv"
        return response

    @staticmethod
    def export_excel(data: List[Dict[str, Any]], fields: List[str], field_labels: Dict[str, str]):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        headers = [field_labels.get(f, f) for f in fields]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            
        for row_num, row_data in enumerate(data, 2):
            for col_num, field in enumerate(fields, 1):
                val = row_data.get(field, "")
                ws.cell(row=row_num, column=col_num, value=str(val))
                
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50)
            
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        response = StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers["Content-Disposition"] = "attachment; filename=report.xlsx"
        return response

    @staticmethod
    def export_pdf(data: List[Dict[str, Any]], fields: List[str], field_labels: Dict[str, str], title: str = "Reporte", current_user: Any = None, metadata: Dict[str, Any] = None):
        stream = io.BytesIO()
        doc = SimpleDocTemplate(stream, pagesize=landscape(letter), rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        elements = []
        
        styles = getSampleStyleSheet()
        brand_color = colors.HexColor('#004562')
        header_bg = colors.HexColor('#004562')
        row_alt_bg = colors.HexColor('#f2f2f2')
        
        # --- Header Section ---
        # Logo Simulation
        logo_text = f'<font name="Times-Italic" size="26" color="{brand_color}">Bienestar</font> <font name="Helvetica" size="20" color="{brand_color}">universitario</font>'
        elements.append(Paragraph(logo_text, styles['Normal']))
        elements.append(Spacer(1, 15))
        
        # Report Details
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        user_name = f"{current_user.first_name} {current_user.last_name}" if current_user else "Administrador"
        
        details_style = ParagraphStyle(
            'Details',
            parent=styles['Normal'],
            fontSize=10,
            leading=14,
            textColor=colors.grey
        )
        
        elements.append(Paragraph(f"<b>Reporte:</b> {title}", styles['Heading2']))
        elements.append(Paragraph(f"<b>Generado por:</b> {user_name}", details_style))
        elements.append(Paragraph(f"<b>Fecha de emisión:</b> {now_str}", details_style))
        
        # Extra info (metadata)
        if metadata:
            elements.append(Spacer(1, 10))
            elements.append(Paragraph("<b>Información Adicional:</b>", styles['Normal']))
            for label, value in metadata.items():
                elements.append(Paragraph(f"• <b>{label}:</b> {value}", details_style))

        elements.append(Spacer(1, 20))
        
        # --- Data Pre-processing (Merge columns if both parts are selected) ---
        processed_fields = []
        processed_labels = {}
        
        # Priority for ID column if selected
        if "id" in fields:
            processed_fields.append("id")
            processed_labels["id"] = field_labels.get("id", "ID")
            
        # Helper to check if both exist in original fields
        has_both_names = "first_name" in fields and "last_name" in fields
        has_both_contact = "email" in fields and "phone" in fields
        
        skip_fields = set(["id"]) # ID handled first
        if has_both_names:
            processed_fields.append("full_name")
            processed_labels["full_name"] = "Nombre Completo"
            skip_fields.update(["first_name", "last_name"])
        
        if has_both_contact:
            processed_fields.append("contact_info")
            processed_labels["contact_info"] = "Contacto"
            skip_fields.update(["email", "phone"])
            
        for f in fields:
            if f not in skip_fields:
                processed_fields.append(f)
                processed_labels[f] = field_labels.get(f, f)
        
        headers = [processed_labels.get(f, f) for f in processed_fields]
        table_data = [headers]
        
        # Mapping for status display
        status_map = {
            "UserStatus.ACTIVE": "Activo",
            "UserStatus.INACTIVE": "Bloqueado",
            "ACTIVE": "Activo",
            "INACTIVE": "Bloqueado",
            "PUBLISHED": "Publicada",
            "FINISHED": "Finalizada",
            "DRAFT": "Borrador",
            "SOLICITADO": "Solicitado",
            "ACTIVO": "Activo",
            "VENCIDO": "Vencido",
            "DEVUELTO": "Devuelto",
            "RECHAZADO": "Rechazado"
        }
        
        for row in data:
            table_row = []
            for field in processed_fields:
                if field == "full_name":
                    val = f"{row.get('first_name', '')} {row.get('last_name', '')}".strip()
                elif field == "contact_info":
                    email = row.get('email', '')
                    phone = row.get('phone', '')
                    val = f"{email}\n{phone}".strip() if email and phone else (email or phone or "")
                else:
                    val = str(row.get(field, ""))
                    # Map status string to friendly name
                    if field == "status" or field == "activity_status":
                        val = status_map.get(val, val)
                    
                    # Format dates (YYYY-MM-DD)
                    if val and any(word in field for word in ["_at", "time", "date"]):
                        if "T" in val:
                            val = val.split("T")[0]
                        elif " " in val:
                            val = val.split(" ")[0]
                    
                    # Format hours (decimal to HH:MM)
                    if "hours" in field:
                        try:
                            f_hours = float(val)
                            h = int(f_hours)
                            m = int(round((f_hours - h) * 60))
                            val = f"{h:02d}:{m:02d}"
                        except (ValueError, TypeError):
                            pass
                
                # Truncate if very long
                if len(val) > 65:
                    val = val[:62] + "..."
                table_row.append(val)
            table_data.append(table_row)
            
        # --- Improved Column Width Allocation ---
        # Define relative weights for columns
        weights = {
            "id": 0.5,
            "full_name": 1.5,
            "first_name": 1.0,
            "last_name": 1.0,
            "national_id": 0.8,
            "contact_info": 1.8,
            "email": 1.2,
            "phone": 0.8,
            "program_name": 1.5,
            "activity_title": 1.5,
            "status": 0.7,
            "created_at": 1.0,
            "scanned_at": 1.0,
            "start_time": 1.0,
            "returned_time": 1.0,
            "hours_earned": 0.6,
            "social_hours_completed": 0.6,
            "issuer_name": 1.0
        }
        
        num_cols = len(processed_fields)
        page_width = landscape(letter)[0] - 80
        
        total_weight = sum(weights.get(f, 1.0) for f in processed_fields)
        col_widths = [(weights.get(f, 1.0) / total_weight) * page_width for f in processed_fields]
        
        t = Table(table_data, colWidths=col_widths)
        t.setStyle(TableStyle([
            # Header Style
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Row Styles
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8), # Slightly smaller for more data
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, row_alt_bg]),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, brand_color),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        def add_footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 8)
            canvas.drawCentredString(landscape(letter)[0]/2, 25, f"Página {canvas.getPageNumber()}")
            canvas.restoreState()

        elements.append(t)
        doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
        
        stream.seek(0)
        
        response = StreamingResponse(stream, media_type="application/pdf")
        response.headers["Content-Disposition"] = f"attachment; filename=reporte_{datetime.now().strftime('%Y%m%d')}.pdf"
        return response
